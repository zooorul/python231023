import openpyxl

# Create a new workbook and select the default worksheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Add headers to the worksheet
sheet['A1'] = "ID"
sheet['B1'] = "Name"
sheet['C1'] = "Quantity"
sheet['D1'] = "Price"

# Generate sample sales data (100 rows)
for i in range(1, 101):
    sheet.cell(row=i + 1, column=1, value=i)  # ID
    sheet.cell(row=i + 1, column=2, value=f"Product {i}")  # Name
    sheet.cell(row=i + 1, column=3, value=i * 2)  # Quantity
    sheet.cell(row=i + 1, column=4, value=i * 10)  # Price

# Save the workbook to the desired location
file_path = "c:\\work\\sales.xlsx"
workbook.save(file_path)

# Close the workbook
workbook.close()

print(f"Sales data saved to {file_path}")
